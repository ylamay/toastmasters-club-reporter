from manager.environment_manager import EnvironmentManager
import pandas as pd
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter # type: ignore
from reportlab.platypus import SimpleDocTemplate # type: ignore
import logging
import json
import os

class FileManager:
    """
    Manages file operations for the Toastmasters data service.

    Attributes:
        env_manager (EnvironmentManager): The environment manager instance
        logger (logging.Logger): The logger instance for this class
    """
    def __init__(self, env_manager: EnvironmentManager):
        self.env_manager = env_manager
        self.logger = logging.getLogger(__name__)

    def save_json(self, data: dict, filename: str, target_dir_attr: str = None) -> bool:
        """
        Save a dictionary as a JSON file to a target location

        Args:
            data (dict): Data to save
            filename (str): Name of the file to save
            target_dir_attr (str): Optional attribute name for a sub-directory within the session directory

        Returns:
            bool: True if save was successful, False otherwise
        """
        try:
            # Validate extension
            if not filename.endswith('.json'):
                self.logger.error("Filename must end with .json")
                return False
            
            # Determine the target directory
            file_path = self._construct_target_path(filename, target_dir_attr)

            # Write the data to the file
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)

            self.logger.info(f"Data written to {file_path}")
            return True
        
        except Exception as e:
            self.logger.error(f"Failed to write data to {filename}: {e}")
            return False

    def load_json(self, filename: str, target_dir_attr: str = None) -> dict:
        """
        Load a JSON file from a target location

        Args:
            filename (str): Name of the file to load
            target_dir_attr (str): Optional attribute name for a sub-directory within the session directory

        Returns:
            dict: Loaded data as a dictionary, or an empty dict if loading failed
        """
        try:
            # Determine the target directory
            file_path = self._construct_target_path(filename, target_dir_attr)

            # Read the data from the file
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            self.logger.info(f"Data loaded from {file_path}")
            return data

        except Exception as e:
            self.logger.warning(f"Failed to load data from {filename}: {e}")
            return {}
        
    def save_markdown(self, content: str, filename: str, target_dir_attr: str = None) -> bool:
        """
        Save a Markdown file to a target location

        Args:
            content (str): Markdown content to save
            filename (str): Name of the file to save
            target_dir_attr (str): Optional attribute name for a sub-directory within the session directory

        Returns:
            bool: True if save was successful, False otherwise
        """
        try:
            # Validate extension
            if not filename.endswith('.md'):
                self.logger.error("Filename must end with .md")
                return False

            # Determine the target directory
            file_path = self._construct_target_path(filename, target_dir_attr)

            # Write the data to the file
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)

            self.logger.info(f"Markdown written to {file_path}")
            return True

        except Exception as e:
            self.logger.error(f"Failed to write Markdown to {filename}: {e}")
            return False
        
    def save_html(self, content: str, filename: str, target_dir_attr: str = None) -> bool:
        """
        Save a HTML file to a target location

        Args:
            content (str): HTML content to save
            filename (str): Name of the file to save
            target_dir_attr (str): Optional attribute name for a sub-directory within the session directory

        Returns:
            bool: True if save was successful, False otherwise
        """
        try:
            # Validate extension
            if not filename.endswith('.html'):
                self.logger.error("Filename must end with .html")
                return False

            # Determine the target directory
            file_path = self._construct_target_path(filename, target_dir_attr)

            # Write the data to the file
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)

            self.logger.info(f"HTML written to {file_path}")
            return True

        except Exception as e:
            self.logger.error(f"Failed to write HTML to {filename}: {e}")
            return False

    def save_excel(self, dataframes: dict, filename: str, target_dir_attr: str = None) -> bool:
        """
        Save multiple dataframes as an Excel file with multiple sheets

        Args:
            dataframes (dict): Dictionary where keys are sheet names and values are pandas DataFrames
            filename (str): Name of the file to save
            target_dir_attr (str): Optional attribute name for a sub-directory within the session directory

        Returns:
            bool: True if save was successful, False otherwise
        """
        try:
            # Validate extension
            if not filename.endswith('.xlsx'):
                self.logger.error("Filename must end with .xlsx")
                return False

            # Determine the target directory
            file_path = self._construct_target_path(filename, target_dir_attr)

            # Write dataframes to Excel file
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in dataframes.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Apply formatting to the worksheet
                    worksheet = writer.sheets[sheet_name]
                    
                    # Format percentage columns
                    self._format_excel_worksheet(worksheet, df)                    

            self.logger.info(f"Excel written to {file_path}")
            return True            

        except Exception as e:
            self.logger.error(f"Failed to write excel file {filename}: {e}")
            return False
        
    def _format_excel_worksheet(self, worksheet, df):
        """
        Apply formatting to Excel worksheet (if applicable)
        
        Args:
            worksheet: openpyxl worksheet object
            df: pandas DataFrame corresponding to the worksheet
        """
        # Find Progress column if it exists
        progress_col = None
        for idx, col_name in enumerate(df.columns):
            if 'Progress' in str(col_name):
                progress_col = idx + 1  # Excel columns are 1-indexed
                break
        
        # Apply percentage formatting to Progress column
        if progress_col:
            for row in range(2, len(df) + 2):  # Skip header row
                cell = worksheet.cell(row=row, column=progress_col)
                if isinstance(cell.value, (int, float)) and cell.value != '':
                    cell.number_format = '0.0%'

        # Bold headers (first row)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        
        # Bold section headers and highlight column headers under Member Pathway Details
        member_pathway_details_row = None
        for row in worksheet.iter_rows():
            first_cell = row[0]
            if first_cell.value and isinstance(first_cell.value, str):
                if any(header in first_cell.value.upper() for header in 
                      ['CLUB OVERVIEW', 'PATHWAY DISTRIBUTION', 'LEVEL DISTRIBUTION', 'MEMBER PATHWAY']):
                    first_cell.font = Font(bold=True, size=12)
                    first_cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

                    # Track the row number for MEMBER PATHWAY DETAILS
                    if 'MEMBER PATHWAY' in first_cell.value.upper():
                        member_pathway_details_row = first_cell.row              

        # Highlight column sub-headers under Member Pathway Details section
        if member_pathway_details_row:
            header_row_number = member_pathway_details_row + 1
            for cell in worksheet[header_row_number]:
                if cell.value:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9EDF7', end_color='D9EDF7', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')     

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    self.logger.warning(f"Error adjusting column width for {column_letter}: {e}")

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def save_pdf(self, elements: list, filename: str, target_dir_attr: str = None) -> bool:
        """
        Save a list of elements as a PDF file
        
        Args:
            elements (list): List of elements to include in the PDF
            filename (str): Name of the file to save
            target_dir_attr (str): Optional attribute name for a sub-directory within the session directory

        Returns:
            bool: True if save was successful, False otherwise
        """
        try:
            # Validate extension
            if not filename.endswith('.pdf'):
                self.logger.error("Filename must end with .pdf")
                return False
            
            # Determine the target directory
            file_path = self._construct_target_path(filename, target_dir_attr)

            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter)

            # Build PDF with the provided elements
            doc.build(elements)

            self.logger.info(f"PDF written to {file_path}")
            return True            

        except Exception as e:
            self.logger.error(f"Failed to write pdf file {filename}: {e}")
            return False

    def _construct_target_path(self, filename: str, target_dir_attr: str) -> str:
        """
        Construct the full file path for saving files
        
        Args:
            filename (str): Name of the file to save
            target_dir_attr (str): Optional attribute name for a sub-directory within the session directory
        
        Returns:
            str: Full file path
        """
        if target_dir_attr:
            target_dir = getattr(self.env_manager, target_dir_attr, None)
            if not target_dir:
                self.logger.error(f"Invalid target directory attribute: {target_dir_attr}")
                return None
        else:
            target_dir = self.env_manager.session_directory

        return os.path.join(target_dir, filename)
